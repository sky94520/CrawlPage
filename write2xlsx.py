import os
import json
import time
import operator
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


basedir = os.path.realpath(os.path.dirname(__file__))
csv_path = os.path.join(basedir, 'files', 'csv')


def main():
    path = os.path.join(basedir, 'files', 'page_links')
    # 遍历整个page_links文件夹
    for parent, dirnames, filenames in os.walk(path, followlinks=True):
        if len(filenames) == 0:
            continue
        # 公司名称
        company = os.path.split(parent)[-1]
        data = []
        # 遍历所有的文件
        for filename in filenames:
            full_filename = os.path.join(parent, filename)
            print(full_filename)
            fp = open(full_filename, 'r', encoding='utf-8')
            json_data = json.load(fp)
            fp.close()
            for datum in json_data:
                del datum['dbcode']
                del datum['dbname']
                del datum['applicants']
                datum['date'] = time.strptime(datum['application_number'], '%Y-%m-%d')

            data.extend(json_data)
        # 筛选并排序
        data = list(filter(lambda x: 2015 <= x['date'].tm_year <= 2019, data))
        data.sort(key=operator.itemgetter('date'), reverse=True)
        # 按照公开号分为发明专利和 实用新型、外观专利
        mapping = {}
        for datum in data:
            publication_number = datum['filename']
            number = publication_number[2]
            if number == '3':
                number = '2'
            if number not in mapping:
                mapping[number] = []
            mapping[number].append(datum)
        # 写入到文件
        real_path = os.path.join(csv_path, '%s.xlsx' % company)

        work_book = Workbook()
        sheet = work_book.create_sheet(index=0)
        i = 1
        fieldnames = ['filename', 'title', 'inventor', 'application_number', 'publication_number']
        keys = ['公开号', '名称', '发明人', '申请日期', '公开日期']
        for j in range(len(keys)):
            sheet.cell(i, j + 1).value = keys[j]
        i += 1
        max_widths = {}
        for number, data in mapping.items():
            type_name = '发明专利' if number == '1' else '其他专利'
            sheet.cell(i, 1).value = '%s个数%d' % (type_name, len(data))
            i += 1
            for datum in data:
                for j in range(len(fieldnames)):
                    value = datum[fieldnames[j]]
                    sheet.cell(i, j + 1).value = value
                    # 确定行的最大值
                    max_widths[j + 1] = max(max_widths.get(j + 1, 0), len(value))
                i += 1
            for k, width in max_widths.items():
                col_letter = get_column_letter(k)
                sheet.column_dimensions[col_letter].width = width * 1.2
            work_book.save(real_path)

        print('%s 写入成功' % company)


if __name__ == '__main__':
    main()
